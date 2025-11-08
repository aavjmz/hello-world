"""
Data Exporter Utility

Exports CAN messages to various formats (CSV, Excel, JSON)
"""

import csv
import json
from pathlib import Path
from typing import List, Optional, Dict, Any
from parsers.asc_parser import CANMessage
from utils.timestamp_formatter import TimestampFormatter, TimestampFormat
from utils.signal_decoder import SignalDecoder


class DataExporter:
    """Exports CAN messages to various file formats"""

    def __init__(self, timestamp_formatter: Optional[TimestampFormatter] = None,
                 signal_decoder: Optional[SignalDecoder] = None):
        """
        Initialize data exporter

        Args:
            timestamp_formatter: Formatter for timestamps (default: RAW format)
            signal_decoder: Signal decoder for decoding messages (optional)
        """
        self.timestamp_formatter = timestamp_formatter or TimestampFormatter(TimestampFormat.RAW)
        self.signal_decoder = signal_decoder

    def export_to_csv(self, messages: List[CANMessage], file_path: str,
                      include_signals: bool = True) -> bool:
        """
        Export messages to CSV format

        Args:
            messages: List of CAN messages
            file_path: Output file path
            include_signals: Whether to include signal values column

        Returns:
            True if export succeeded, False otherwise
        """
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                # Define headers
                headers = ['序号', '时间戳', '通道', 'CAN ID', '方向', 'DLC', '数据']
                if include_signals and self.signal_decoder:
                    headers.append('信号值')

                writer = csv.writer(csvfile)
                writer.writerow(headers)

                # Write message data
                for idx, message in enumerate(messages, start=1):
                    row = [
                        idx,
                        self.timestamp_formatter.format(message.timestamp),
                        message.channel,
                        f'0x{message.can_id:03X}',
                        message.direction,
                        message.dlc,
                        ' '.join(f'{b:02X}' for b in message.data)
                    ]

                    # Add signal values if requested
                    if include_signals and self.signal_decoder:
                        decoded = self.signal_decoder.decode_message(message)
                        if decoded and decoded.signals:
                            signal_str = ', '.join(
                                f'{name}: {value.physical_value}{value.unit}'
                                for name, value in decoded.signals.items()
                            )
                            row.append(signal_str)
                        else:
                            row.append('')

                    writer.writerow(row)

            return True

        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False

    def export_to_excel(self, messages: List[CANMessage], file_path: str,
                       include_signals: bool = True) -> bool:
        """
        Export messages to Excel format

        Args:
            messages: List of CAN messages
            file_path: Output file path
            include_signals: Whether to include signal values column

        Returns:
            True if export succeeded, False otherwise
        """
        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                print("Error: openpyxl library not installed. Please install it with: pip install openpyxl")
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = "CAN Messages"

            # Define headers
            headers = ['序号', '时间戳', '通道', 'CAN ID', '方向', 'DLC', '数据']
            if include_signals and self.signal_decoder:
                headers.append('信号值')

            # Write headers with formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Write message data
            for row_idx, message in enumerate(messages, start=2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
                ws.cell(row=row_idx, column=2, value=self.timestamp_formatter.format(message.timestamp))
                ws.cell(row=row_idx, column=3, value=message.channel)
                ws.cell(row=row_idx, column=4, value=f'0x{message.can_id:03X}')
                ws.cell(row=row_idx, column=5, value=message.direction)
                ws.cell(row=row_idx, column=6, value=message.dlc)
                ws.cell(row=row_idx, column=7, value=' '.join(f'{b:02X}' for b in message.data))

                # Add signal values if requested
                if include_signals and self.signal_decoder:
                    decoded = self.signal_decoder.decode_message(message)
                    if decoded and decoded.signals:
                        signal_str = ', '.join(
                            f'{name}: {value.physical_value}{value.unit}'
                            for name, value in decoded.signals.items()
                        )
                        ws.cell(row=row_idx, column=8, value=signal_str)
                    else:
                        ws.cell(row=row_idx, column=8, value='')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False

    def export_to_json(self, messages: List[CANMessage], file_path: str,
                      include_signals: bool = True, pretty: bool = True) -> bool:
        """
        Export messages to JSON format

        Args:
            messages: List of CAN messages
            file_path: Output file path
            include_signals: Whether to include decoded signal values
            pretty: Whether to format JSON with indentation

        Returns:
            True if export succeeded, False otherwise
        """
        try:
            # Build JSON structure
            data = {
                'metadata': {
                    'total_messages': len(messages),
                    'timestamp_format': self.timestamp_formatter.format_type.value,
                    'has_signal_data': include_signals and self.signal_decoder is not None
                },
                'messages': []
            }

            # Add message data
            for message in messages:
                msg_dict = {
                    'timestamp': message.timestamp,
                    'timestamp_formatted': self.timestamp_formatter.format(message.timestamp),
                    'channel': message.channel,
                    'can_id': f'0x{message.can_id:03X}',
                    'can_id_decimal': message.can_id,
                    'direction': message.direction,
                    'dlc': message.dlc,
                    'data_hex': ' '.join(f'{b:02X}' for b in message.data),
                    'data_bytes': list(message.data)
                }

                # Add signal values if requested
                if include_signals and self.signal_decoder:
                    decoded = self.signal_decoder.decode_message(message)
                    if decoded and decoded.signals:
                        msg_dict['signals'] = {
                            name: {
                                'physical_value': value.physical_value,
                                'raw_value': value.raw_value,
                                'unit': value.unit
                            }
                            for name, value in decoded.signals.items()
                        }
                    else:
                        msg_dict['signals'] = {}

                data['messages'].append(msg_dict)

            # Write JSON file
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                if pretty:
                    json.dump(data, jsonfile, indent=2, ensure_ascii=False)
                else:
                    json.dump(data, jsonfile, ensure_ascii=False)

            return True

        except Exception as e:
            print(f"Error exporting to JSON: {e}")
            return False

    def get_export_statistics(self, messages: List[CANMessage]) -> Dict[str, Any]:
        """
        Get statistics about messages to be exported

        Args:
            messages: List of CAN messages

        Returns:
            Dictionary containing statistics
        """
        if not messages:
            return {
                'total_messages': 0,
                'unique_ids': 0,
                'rx_count': 0,
                'tx_count': 0,
                'time_span': 0.0
            }

        unique_ids = set(msg.can_id for msg in messages)
        rx_count = sum(1 for msg in messages if msg.direction == 'Rx')
        tx_count = sum(1 for msg in messages if msg.direction == 'Tx')

        timestamps = [msg.timestamp for msg in messages]
        time_span = max(timestamps) - min(timestamps) if timestamps else 0.0

        return {
            'total_messages': len(messages),
            'unique_ids': len(unique_ids),
            'rx_count': rx_count,
            'tx_count': tx_count,
            'time_span': time_span,
            'start_time': min(timestamps) if timestamps else 0.0,
            'end_time': max(timestamps) if timestamps else 0.0
        }
